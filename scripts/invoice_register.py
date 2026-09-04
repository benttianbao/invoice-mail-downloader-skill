#!/usr/bin/env python3
"""用 openpyxl 生成可移植的发票登记表。"""

from __future__ import annotations

import datetime as dt
from pathlib import Path
import tempfile
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


SHEET_NAME = "发票登记"
HEADERS = [
    "序号",
    "发票日期",
    "支付日期",
    "收款姓名",
    "发票金额",
    "报销金额",
    "费用类型",
    "开票方",
    "发票编号",
    "发票下载时间",
    "备注",
    "是否存在报销凭证",
    "报销凭证",
    "凭证匹配状态",
    "凭证校验说明",
    "发票与实付差额",
    "校验状态",
]
EXPENSE_TYPES = ("部门营销费用", "企业文化费用", "出差报销费用")
AUTO_FILL = PatternFill("solid", fgColor="EAF4F8")
MANUAL_FILL = PatternFill("solid", fgColor="FFF7D6")
HEADER_FILL = PatternFill("solid", fgColor="319795")
TITLE_FILL = PatternFill("solid", fgColor="234E52")
NOTE_FILL = PatternFill("solid", fgColor="E6FFFA")
THIN = Border(
    left=Side(style="thin", color="D7E0E5"),
    right=Side(style="thin", color="D7E0E5"),
    top=Side(style="thin", color="D7E0E5"),
    bottom=Side(style="thin", color="D7E0E5"),
)
WIDTHS = {
    "A": 8,
    "B": 13,
    "C": 13,
    "D": 15,
    "E": 14,
    "F": 14,
    "G": 18,
    "H": 40,
    "I": 24,
    "J": 20,
    "K": 42,
    "L": 16,
    "M": 48,
    "N": 18,
    "O": 52,
    "P": 18,
    "Q": 30,
    "R": 2,
}


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    return str(value).strip()


def read_manual_values(path: Path) -> dict[str, dict[str, Any]]:
    workbook = load_workbook(path, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        return {}
    sheet = workbook[SHEET_NAME]
    header_index = None
    header: list[Any] = []
    for index, row in enumerate(sheet.iter_rows(min_row=1, max_row=8, values_only=True), start=1):
        values = list(row or [])
        if values and values[0] == "序号" and len(values) > 8 and values[8] == "发票编号":
            header_index = index
            header = values
            break
    if header_index is None:
        return {}

    def index_of(name: str, fallback: int) -> int:
        try:
            return header.index(name)
        except ValueError:
            return fallback

    columns = {
        "paymentDate": index_of("支付日期", 2),
        "recipient": index_of("收款姓名", 3),
        "reimbursementAmount": index_of("报销金额", 5),
        "expenseType": index_of("费用类型", 6),
        "note": index_of("备注", 10),
    }
    key_column = header.index("校验状态") + 1 if "校验状态" in header else 12
    saved: dict[str, dict[str, Any]] = {}
    for row in sheet.iter_rows(min_row=header_index + 1, values_only=True):
        values = list(row or [])
        if key_column >= len(values):
            continue
        key = _cell_text(values[key_column])
        if not key:
            continue
        saved[key] = {
            "paymentDate": values[columns["paymentDate"]] if columns["paymentDate"] < len(values) else None,
            "recipient": values[columns["recipient"]] if columns["recipient"] < len(values) else "",
            "reimbursementAmount": values[columns["reimbursementAmount"]] if columns["reimbursementAmount"] < len(values) else None,
            "expenseType": values[columns["expenseType"]] if columns["expenseType"] < len(values) else "",
            "note": values[columns["note"]] if columns["note"] < len(values) else "",
        }
    return saved


def _date_only(value: Any) -> dt.date | str | None:
    text = _cell_text(value)
    if not text:
        return None
    try:
        return dt.date.fromisoformat(text[:10])
    except ValueError:
        return text


def _date_time(value: Any) -> dt.datetime | str | None:
    text = _cell_text(value)
    if not text:
        return None
    try:
        parsed = dt.datetime.fromisoformat(text)
        if parsed.tzinfo is not None:
            parsed = parsed.astimezone().replace(tzinfo=None)
        return parsed
    except ValueError:
        return text


def _amount(value: Any) -> float | str | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return value


def write_invoice_workbook(output: Path, records: list[dict[str, Any]]) -> dict[str, Any]:
    saved: dict[str, dict[str, Any]] = {}
    if output.exists():
        try:
            saved = read_manual_values(output)
        except Exception as exc:  # noqa: BLE001 - 保护人工字段，原样上报
            raise RuntimeError(f"现有发票登记表无法读取，为保护人工数据未覆盖：{exc}") from exc

    records = sorted(
        records,
        key=lambda record: (
            str(record.get("downloaded_at") or ""),
            str(record.get("invoice_date") or ""),
            str(record.get("record_key") or ""),
        ),
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A4"
    sheet.merge_cells("A1:Q1")
    sheet.merge_cells("A2:Q2")
    sheet["A1"] = "发票登记表"
    sheet["A1"].font = Font(bold=True, color="FFFFFF", size=18)
    sheet["A1"].fill = TITLE_FILL
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet["A2"] = "浅蓝色为自动填写字段；浅黄色为人工填写字段；凭证由 TextIn XParse 公有服务识别；敏感凭证请勿放入投放目录。"
    sheet["A2"].font = Font(italic=True, color="285E61", size=10)
    sheet["A2"].fill = NOTE_FILL
    sheet["A2"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 34
    sheet.row_dimensions[2].height = 24
    sheet.row_dimensions[3].height = 30
    for index, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(3, index, header)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN
    sheet.cell(3, 18, "")

    auto_columns = {1, 2, 5, 8, 9, 10, 12, 13, 14, 15, 16, 17}
    manual_columns = {3, 4, 6, 7, 11}
    for offset, record in enumerate(records):
        row_number = offset + 4
        manual = saved.get(str(record.get("record_key") or ""), {})
        reimbursement = manual.get("reimbursementAmount")
        if reimbursement in (None, ""):
            reimbursement = _amount(record.get("invoice_amount"))
        payment = manual.get("paymentDate") or _date_only(record.get("receipt_payment_date"))
        values = [
            offset + 1,
            _date_only(record.get("invoice_date")),
            payment,
            manual.get("recipient") or "",
            _amount(record.get("invoice_amount")),
            reimbursement,
            manual.get("expenseType") or "",
            record.get("seller") or "",
            record.get("invoice_number") or "",
            _date_time(record.get("downloaded_at")),
            manual.get("note") or "",
            record.get("has_receipt") or "否",
            record.get("receipt_path") or "",
            record.get("receipt_status") or "未匹配",
            record.get("receipt_detail") or "尚未找到对应报销凭证",
            _amount(record.get("receipt_amount_difference")),
            record.get("validation_status") or "待确认",
            record.get("record_key") or "",
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_number, column, value)
            cell.alignment = Alignment(vertical="center", wrap_text=column in {8, 11, 13, 15, 17})
            cell.border = THIN
            if column in auto_columns:
                cell.fill = AUTO_FILL
            elif column in manual_columns:
                cell.fill = MANUAL_FILL
        sheet.row_dimensions[row_number].height = 32
        sheet.cell(row_number, 2).number_format = "yyyy/m/d"
        sheet.cell(row_number, 3).number_format = "yyyy/m/d"
        sheet.cell(row_number, 5).number_format = '¥#,##0.00'
        sheet.cell(row_number, 6).number_format = '¥#,##0.00'
        sheet.cell(row_number, 16).number_format = '¥#,##0.00'
        sheet.cell(row_number, 9).number_format = "@"
        sheet.cell(row_number, 10).number_format = "yyyy/m/d hh:mm"
        sheet.cell(row_number, 18).font = Font(color="FFFFFF", size=1)

    last_data_row = max(4, len(records) + 3)
    if records:
        table = Table(displayName="InvoiceRegisterTable", ref=f"A3:Q{last_data_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
        sheet.add_table(table)
        sheet.conditional_formatting.add(
            f"N4:N{last_data_row}",
            FormulaRule(
                formula=['OR($N4="未匹配",LEFT($N4,4)="匹配冲突",LEFT($N4,4)="识别失败",LEFT($N4,3)="待确认")'],
                fill=PatternFill("solid", fgColor="FED7D7"),
                font=Font(color="9B2C2C", bold=True),
            ),
        )
        sheet.conditional_formatting.add(
            f"N4:N{last_data_row}",
            FormulaRule(
                formula=['OR($N4="已匹配",$N4="人工匹配")'],
                fill=PatternFill("solid", fgColor="C6F6D5"),
                font=Font(color="22543D", bold=True),
            ),
        )
        sheet.conditional_formatting.add(
            f"Q4:Q{last_data_row}",
            FormulaRule(
                formula=['LEFT($Q4,3)="待确认"'],
                fill=PatternFill("solid", fgColor="FED7D7"),
                font=Font(color="9B2C2C", bold=True),
            ),
        )
        sheet.conditional_formatting.add(
            f"Q4:Q{last_data_row}",
            FormulaRule(
                formula=['$Q4="通过"'],
                fill=PatternFill("solid", fgColor="C6F6D5"),
                font=Font(color="22543D", bold=True),
            ),
        )

    validation_last_row = max(200, last_data_row + 50)
    validation = DataValidation(type="list", formula1='"' + ",".join(EXPENSE_TYPES) + '"', allow_blank=True)
    validation.add(f"G4:G{validation_last_row}")
    sheet.add_data_validation(validation)
    for column, width in WIDTHS.items():
        sheet.column_dimensions[column].width = width
    sheet.column_dimensions["R"].hidden = True

    output.parent.mkdir(parents=True, exist_ok=True)
    handle = tempfile.NamedTemporaryFile(prefix=output.name + ".", suffix=".xlsx", dir=output.parent, delete=False)
    temp_path = Path(handle.name)
    handle.close()
    try:
        workbook.save(temp_path)
        temp_path.replace(output)
    finally:
        temp_path.unlink(missing_ok=True)
    return {"status": "success", "path": str(output), "rows": len(records), "backend": "openpyxl"}
